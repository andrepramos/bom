from anytree import NodeMixin, find_by_attr, RenderTree, PreOrderIter
import argparse
from glob import glob
import csv
from openpyxl import Workbook
import re
import sys
from contextlib import suppress

BAD_PN_REGEX = [r'^$', r'^X\d+_[A-Z]', r'\d+_SUB\d+']


def _tuppled_item(item_str: str) -> tuple:
    if item_str:
        return tuple([int(item) for item in item_str.split('.')])
    else:
        return ()


def valid_files():
    files = glob('*.txt')
    with suppress(ValueError):
        files.remove('ignore.txt')
    return files


class ItemNumber:
    def __init__(self, item: str):
        self.tupple = _tuppled_item(item)

    def get_parent(self):
        parent_string = '.'.join(self.__repr__().split('.')[:-1])
        return ItemNumber(parent_string)

    def __repr__(self):
        return '.'.join(str(num) for num in self.tupple)

    def __eq__(self, other):
        return self.tupple == other.tupple

    @staticmethod
    def _filter_pn(part_number: str) -> str:
        re_lst = []
        for regex in BAD_PN_REGEX:
            re_lst.append(re.compile(regex))

        if any(regex.search(part_number) for regex in re_lst):
            raise ValueError("Part number matches invalid regex")
        else:
            return part_number


class BOMItem(NodeMixin):
    def __init__(self,
                 item_n: ItemNumber,
                 part_n: str,
                 descript: str,
                 qty: str,
                 parent=None):
        super(BOMItem, self).__init__()
        self.item_n = ItemNumber(item_n)
        self.part_n = ItemNumber._filter_pn(part_n)
        self.descript = descript
        self.qty = int(qty)
        self.parent = parent

    def __repr__(self):
        return f"{self.item_n} -- {self.part_n} -- {self.qty}"

    def __eq__(self, other):
        return self.part_n == other.part_n


class BOM:
    def __init__(self, filename):
        self.tree = BOMItem('', 'Part Number', 'Description', '1')
        self.top_item = filename.split('.')[0]
        self.ignorelist = BOM.read_ignorefile('ignore.txt')
        with open(filename) as csvfile:
            bomreader = csv.reader(csvfile, delimiter='\t')
            next(bomreader)
            for row in bomreader:
                if row[1].strip() in self.ignorelist:
                    continue
                else:
                    with suppress(ValueError, TypeError):
                        parent = self._get_parent(row[0])
                        BOMItem(
                            row[0],
                            row[1].strip(),
                            row[2],
                            row[3],
                            parent=parent)

    def _get_parent(self, item: str) -> BOMItem:
        item = ItemNumber(item)
        return find_by_attr(self.tree, item.get_parent(), "item_n")

    def print(self):
        for pre, _, node in RenderTree(self.tree):
            treestr = u"%s%s" % (pre, node)
            print(treestr.ljust(8))

    def flat(self):
        '''Returns flat list'''
        flat_dict = {}
        for node in PreOrderIter(self.tree):
            qty_sum = 1
            for inner_node in node.path:
                qty_sum *= inner_node.qty

            if (node.part_n, node.descript) in flat_dict.keys():
                flat_dict[(node.part_n, node.descript)] += qty_sum
            else:
                flat_dict[(node.part_n, node.descript)] = qty_sum

        flat_list = []
        for key, value in flat_dict.items():
            flat_list.append((*key, value))

        flat_list[0] = ('Part Number', 'Description', 'Qty.')

        return flat_list

    def indented(self) -> list:
        indented_list = []
        for node in PreOrderIter(self.tree):
            indented_list.append((str(node.item_n), node.part_n, node.descript,
                                  node.qty))
        indented_list[0] = ('Item', 'Part Number', 'Description', 'Qty.')
        return indented_list

    @staticmethod
    def _write_to_sheet(lst, ws):
        for row, bom_item in enumerate(lst):
            excel_row = row + 1
            for col, prop in enumerate(bom_item):
                excel_col = col + 1
                cell = ws.cell(row=excel_row, column=excel_col)
                cell.value = prop

    @staticmethod
    def read_ignorefile(ignorefile):
        try:
            with open(ignorefile, 'r') as f:
                return f.read().splitlines()
        except FileNotFoundError:
            return []

    def write_file(self):
        wb = Workbook()

        ws_indented = wb.active
        ws_indented.title = "Indented"

        ws_flat = wb.create_sheet("Flat")

        BOM._write_to_sheet(self.indented(), ws_indented)
        BOM._write_to_sheet(self.flat(), ws_flat)

        wb.save((self.top_item + '.xlsx'))


if __name__ == '__main__':
    files = valid_files()
    parser = argparse.ArgumentParser()
    parser.add_argument('files', help='Files to write BOM', nargs='*')
    args = parser.parse_args()
    if args.files:
        for file in args.files:
            bom = BOM(file)
            bom.write_file()
    elif files:
        for file in files:
            bom = BOM(file)
            bom.write_file()
    else:
        print("There are no valid files.")